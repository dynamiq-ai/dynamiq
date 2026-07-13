import copy
import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from dynamiq.components.converters.base import BaseConverter
from dynamiq.components.converters.utils import (
    build_source_metadata,
    get_filename_for_bytesio,
    normalize_table_headers,
)
from dynamiq.types import Document, DocumentCreationMode, DocumentType

SUPPORTED_EXCEL_EXTENSIONS = {".csv", ".tsv", ".xlsx"}
DelimitedDocumentCreationMode = Literal[
    DocumentCreationMode.ONE_DOC_PER_FILE,
    DocumentCreationMode.ONE_DOC_PER_ROW,
]
WorkbookDocumentCreationMode = Literal[
    DocumentCreationMode.ONE_DOC_PER_FILE,
    DocumentCreationMode.ONE_DOC_PER_SHEET,
    DocumentCreationMode.ONE_DOC_PER_ROW,
]


class ExcelFileConverter(BaseConverter):
    """
    A component for converting spreadsheet files (xlsx, csv, tsv) to Documents.

    Excel workbooks are read with openpyxl and every sheet is rendered as a markdown
    table (one section per sheet). CSV/TSV files are parsed with the stdlib csv module
    and rendered as a single markdown table.

    Args:
        document_creation_mode (Literal["one-doc-per-file"], optional):
            Determines how to create Documents from the spreadsheet content. Currently only
            supports:
            - `"one-doc-per-file"`: Creates one Document per file.
            Defaults to `"one-doc-per-file"`.

    Usage example:
        ```python
        from dynamiq.components.converters.excel import ExcelFileConverter

        converter = ExcelFileConverter()
        documents = converter.run(file_paths=["a/file/path.xlsx"])["documents"]
        ```
    """

    document_creation_mode: Literal[DocumentCreationMode.ONE_DOC_PER_FILE] = DocumentCreationMode.ONE_DOC_PER_FILE
    delimited_document_creation_mode: DelimitedDocumentCreationMode = DocumentCreationMode.ONE_DOC_PER_FILE
    workbook_document_creation_mode: WorkbookDocumentCreationMode = DocumentCreationMode.ONE_DOC_PER_FILE

    def _process_file(self, file: Path | BytesIO, metadata: dict[str, Any]) -> list[Any]:
        """
        Process a file and return a list of Documents.

        Args:
            file: Path to a file or BytesIO object
            metadata: Metadata to attach to the documents

        Returns:
            List of Documents
        """
        if isinstance(file, BytesIO):
            filepath = get_filename_for_bytesio(file)
            file.seek(0)
            data = file.read()
            file.seek(0)
        else:
            filepath = str(file)
            with open(file, "rb") as f:
                data = f.read()

        extension = Path(filepath).suffix.lower()
        if extension and extension not in SUPPORTED_EXCEL_EXTENSIONS:
            supported_extensions = ", ".join(sorted(SUPPORTED_EXCEL_EXTENSIONS))
            raise ValueError(
                f"Unsupported spreadsheet extension '{extension}'. "
                f"ExcelFileConverter supports: {supported_extensions}."
            )

        if extension in {".csv", ".tsv"}:
            delimiter = "\t" if extension == ".tsv" else ","
            documents = self._create_delimited_documents(data, filepath, metadata, delimiter)
        else:
            delimiter = None if extension else self._detect_delimiter(data)
            if delimiter:
                documents = self._create_delimited_documents(data, filepath, metadata, delimiter)
            else:
                documents = self._create_workbook_documents(data, filepath, metadata)

        if not documents:
            raise ValueError(f"Spreadsheet file '{filepath}' contains no extractable content.")
        return documents

    def _create_delimited_documents(
        self,
        data: bytes,
        filepath: str,
        metadata: dict[str, Any],
        delimiter: str,
    ) -> list[Document]:
        content_type = "text/tab-separated-values" if delimiter == "\t" else "text/csv"
        if self.delimited_document_creation_mode == DocumentCreationMode.ONE_DOC_PER_ROW:
            return self._create_delimited_row_documents(
                filepath=filepath,
                rows=self._read_delimited_rows(data, delimiter=delimiter),
                metadata=metadata,
                content_type=content_type,
            )
        return self._create_documents(
            filepath=filepath,
            content=self._convert_delimited(data, delimiter=delimiter),
            document_creation_mode=self.document_creation_mode,
            metadata=metadata,
        )

    def _create_workbook_documents(
        self,
        data: bytes,
        filepath: str,
        metadata: dict[str, Any],
    ) -> list[Document]:
        creators = {
            DocumentCreationMode.ONE_DOC_PER_ROW: self._create_workbook_row_documents,
            DocumentCreationMode.ONE_DOC_PER_SHEET: self._create_workbook_sheet_documents,
        }
        creator = creators.get(self.workbook_document_creation_mode)
        if creator:
            return creator(data, filepath, metadata)
        return self._create_documents(
            filepath=filepath,
            content=self._convert_workbook(data),
            document_creation_mode=self.document_creation_mode,
            metadata=metadata,
        )

    def _convert_workbook(self, data: bytes) -> str:
        """Convert an Excel workbook to markdown, one table section per sheet."""
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        try:
            sections = []
            for worksheet in workbook.worksheets:
                rows = [
                    ["" if cell is None else str(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)
                ]
                table = self._to_markdown_table(rows)
                if table:
                    sections.append(f"## {worksheet.title}\n\n{table}" if len(workbook.worksheets) > 1 else table)
            return "\n\n".join(sections)
        finally:
            workbook.close()

    def _convert_delimited(self, data: bytes, delimiter: str) -> str:
        """Convert CSV/TSV content to a markdown table."""
        return self._to_markdown_table(self._read_delimited_rows(data, delimiter))

    @classmethod
    def _create_workbook_row_documents(
        cls,
        data: bytes,
        filepath: str,
        metadata: dict[str, Any],
    ) -> list[Document]:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        try:
            documents: list[Document] = []
            for worksheet in workbook.worksheets:
                indexed_rows = [
                    (row_number, ["" if value is None else str(value) for value in row])
                    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1)
                    if any(value is not None and str(value).strip() for value in row)
                ]
                if len(indexed_rows) < 2:
                    continue

                _, raw_headers = indexed_rows[0]
                data_rows = indexed_rows[1:]
                width = max(len(raw_headers), *(len(row) for _, row in data_rows))
                headers = normalize_table_headers(raw_headers, width)
                for row_number, row in data_rows:
                    padded = row + [""] * (width - len(row))
                    fields = [f"{headers[index]}: {value}" for index, value in enumerate(padded) if value.strip()]
                    if not fields:
                        continue
                    row_metadata = cls._spreadsheet_metadata(
                        metadata,
                        filepath,
                        worksheet.title,
                        document_type=DocumentType.TABLE_ROW,
                    )
                    row_metadata["row_number"] = row_number
                    documents.append(
                        Document(content=f"Sheet: {worksheet.title}\n" + "\n".join(fields), metadata=row_metadata)
                    )
            return documents
        finally:
            workbook.close()

    @classmethod
    def _create_workbook_sheet_documents(
        cls,
        data: bytes,
        filepath: str,
        metadata: dict[str, Any],
    ) -> list[Document]:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        try:
            documents: list[Document] = []
            for worksheet in workbook.worksheets:
                rows = [
                    ["" if value is None else str(value) for value in row]
                    for row in worksheet.iter_rows(values_only=True)
                ]
                content = cls._to_markdown_table(rows)
                if not content:
                    continue
                sheet_metadata = cls._spreadsheet_metadata(
                    metadata,
                    filepath,
                    worksheet.title,
                    document_type=DocumentType.TABLE,
                )
                sheet_metadata["row_count"] = max(
                    0, len([row for row in rows if any(cell.strip() for cell in row)]) - 1
                )
                documents.append(Document(content=content, metadata=sheet_metadata))
            return documents
        finally:
            workbook.close()

    @classmethod
    def _spreadsheet_metadata(
        cls,
        metadata: dict[str, Any],
        filepath: str,
        sheet_name: str,
        document_type: DocumentType,
    ) -> dict[str, Any]:
        result = build_source_metadata(metadata, filepath)
        result["sheet_name"] = sheet_name
        result["document_type"] = document_type.value
        result.setdefault("content_type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return result

    @staticmethod
    def _read_delimited_rows(data: bytes, delimiter: str) -> list[list[str]]:
        """Decode a delimited file and return logical rows, including multiline cells."""
        text = data.decode("utf-8-sig", errors="replace")
        return list(csv.reader(StringIO(text), delimiter=delimiter))

    @classmethod
    def _create_delimited_row_documents(
        cls,
        filepath: str,
        rows: list[list[str]],
        metadata: dict[str, Any],
        content_type: str,
    ) -> list[Document]:
        """Create one self-describing document per CSV/TSV row.

        Column labels are repeated in the embedded text so a retrieved row remains
        understandable without its original table or a CSV-specific system prompt.
        """
        indexed_rows = [
            (row_number, row) for row_number, row in enumerate(rows, start=1) if any(cell.strip() for cell in row)
        ]
        if len(indexed_rows) < 2:
            return []

        _, header = indexed_rows[0]
        data_rows = indexed_rows[1:]
        width = max(len(header), *(len(row) for _, row in data_rows))
        headers = normalize_table_headers(header, width)
        source_metadata = build_source_metadata(metadata, filepath)
        source_metadata.setdefault("content_type", content_type)
        source_metadata.setdefault("document_type", DocumentType.TABLE_ROW.value)

        documents: list[Document] = []
        for row_number, row in data_rows:
            padded = row + [""] * (width - len(row))
            fields = [f"{headers[index]}: {value}" for index, value in enumerate(padded) if value.strip()]
            if not fields:
                continue
            row_metadata = copy.deepcopy(source_metadata)
            row_metadata["row_number"] = row_number
            documents.append(Document(content="\n".join(fields), metadata=row_metadata))
        return documents

    @staticmethod
    def _detect_delimiter(data: bytes) -> str | None:
        """Detect extensionless CSV/TSV content without treating binary workbooks as text."""
        if b"\x00" in data[:8192]:
            return None
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            return None

        lines = [line for line in text.splitlines() if line.strip()]
        if len(lines) < 2:
            return None

        sample = "\n".join(lines[:10])
        for delimiter in (",", "\t"):
            rows = list(csv.reader(StringIO(sample), delimiter=delimiter))
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if len(rows) < 2:
                continue

            widths = [len(row) for row in rows]
            if max(widths) < 2:
                continue

            most_common_width = max(set(widths), key=widths.count)
            if most_common_width >= 2 and widths.count(most_common_width) / len(widths) >= 0.8:
                return delimiter

        return None

    @staticmethod
    def _to_markdown_table(rows: list[list[str]]) -> str:
        """Render rows as a markdown table, treating the first row as the header."""
        rows = [row for row in rows if any(cell.strip() for cell in row)]
        if not rows:
            return ""

        width = max(len(row) for row in rows)

        def render_row(row: list[str]) -> str:
            padded = row + [""] * (width - len(row))
            cells = [cell.replace("|", "\\|").replace("\n", " ") for cell in padded]
            return "| " + " | ".join(cells) + " |"

        lines = [render_row(rows[0]), "| " + " | ".join(["---"] * width) + " |"]
        lines.extend(render_row(row) for row in rows[1:])
        return "\n".join(lines)

    def _create_documents(
        self,
        filepath: str,
        content: str,
        document_creation_mode: DocumentCreationMode,
        metadata: dict[str, Any],
        **kwargs,
    ) -> list[Document]:
        """
        Create Documents from the spreadsheet content.
        """
        if document_creation_mode != DocumentCreationMode.ONE_DOC_PER_FILE:
            raise ValueError("ExcelFileConverter only supports one-doc-per-file mode")

        if not content.strip():
            raise ValueError(f"Spreadsheet file '{filepath}' contains no extractable content.")

        metadata = build_source_metadata(metadata, filepath)

        return [Document(content=content.strip(), metadata=metadata)]
